from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

COLOR_VERDE = colors.HexColor("#2B9E6F")
COLOR_VERDE_OSCURO = colors.HexColor("#1f7f56")
COLOR_VERDE_CLARO = colors.HexColor("#e6f4ef")
COLOR_GRIS = colors.HexColor("#666666")
COLOR_GRIS_CLARO = colors.HexColor("#f8f8f8")
COLOR_TEXTO = colors.HexColor("#1a1a2e")
COLOR_BORDE = colors.HexColor("#e8e8e8")
COLOR_BLANCO = colors.white

def generate_pdf_report(employee_id: str, kpis: list, output_path: str):
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch
    )

    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "Titulo",
        parent=styles["Normal"],
        fontSize=22,
        fontName="Helvetica-Bold",
        textColor=COLOR_BLANCO,
        alignment=TA_LEFT,
        spaceAfter=4,
    )
    estilo_subtitulo = ParagraphStyle(
        "Subtitulo",
        parent=styles["Normal"],
        fontSize=11,
        fontName="Helvetica",
        textColor=colors.HexColor("#d0ede2"),
        alignment=TA_LEFT,
    )
    estilo_seccion = ParagraphStyle(
        "Seccion",
        parent=styles["Normal"],
        fontSize=11,
        fontName="Helvetica-Bold",
        textColor=COLOR_VERDE_OSCURO,
        spaceAfter=6,
        spaceBefore=10,
    )
    estilo_footer = ParagraphStyle(
        "Footer",
        parent=styles["Normal"],
        fontSize=9,
        fontName="Helvetica",
        textColor=COLOR_GRIS,
        alignment=TA_CENTER,
    )

    story = []

    ancho_util = letter[0] - 1.5 * inch
    encabezado_data = [[
        Paragraph(f"Reporte de Desempeño", estilo_titulo),
        Paragraph(f"Empleado #{employee_id}", estilo_titulo),
    ]]
    encabezado = Table(encabezado_data, colWidths=[ancho_util * 0.65, ancho_util * 0.35])
    encabezado.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_VERDE),
        ("TOPPADDING", (0, 0), (-1, -1), 22),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 22),
        ("LEFTPADDING", (0, 0), (0, -1), 24),
        ("RIGHTPADDING", (-1, 0), (-1, -1), 24),
        ("ROUNDEDCORNERS", [10, 10, 10, 10]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(encabezado)
    story.append(Spacer(1, 18))

    story.append(HRFlowable(
        width="100%", thickness=1,
        color=COLOR_BORDE, spaceAfter=12
    ))

    story.append(Paragraph("Indicadores de Desempeño (KPIs)", estilo_seccion))

    if kpis:
        data = [["KPI", "Valor Actual", "Meta", "% Logro"]]
        for kpi in kpis:
            porcentaje = kpi.porcentaje
            data.append([
                kpi.kpi_nombre,
                str(kpi.valor_actual),
                str(kpi.valor_meta),
                f"{porcentaje}%"
            ])

        col_widths = [
            ancho_util * 0.40,
            ancho_util * 0.20,
            ancho_util * 0.20,
            ancho_util * 0.20,
        ]
        tabla = Table(data, colWidths=col_widths, repeatRows=1)

        tabla_style = TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), COLOR_VERDE),
            ("TEXTCOLOR", (0, 0), (-1, 0), COLOR_BLANCO),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("ALIGN", (1, 0), (-1, 0), "CENTER"),

            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("TEXTCOLOR", (0, 1), (-1, -1), COLOR_TEXTO),
            ("TOPPADDING", (0, 1), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [COLOR_BLANCO, COLOR_VERDE_CLARO]),

            ("LINEBELOW", (0, 0), (-1, 0), 2, COLOR_VERDE_OSCURO),
            ("LINEBELOW", (0, 1), (-1, -1), 0.5, COLOR_BORDE),
            ("BOX", (0, 0), (-1, -1), 1, COLOR_BORDE),
        ])
        tabla.setStyle(tabla_style)
        story.append(tabla)
    else:
        story.append(Paragraph("No se encontraron KPIs para este empleado.", styles["Normal"]))

    story.append(Spacer(1, 24))
    story.append(HRFlowable(width="100%", thickness=1, color=COLOR_BORDE, spaceAfter=8))
    story.append(Paragraph("Generado por PerformTrack · Módulo para MintHCM", estilo_footer))

    doc.build(story)


def generate_excel_report(employee_id: str, kpis: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluacion de Desempeno"

    VERDE_HEX = "2B9E6F"
    VERDE_OSCURO_HEX = "1f7f56"
    VERDE_CLARO_HEX = "e6f4ef"
    GRIS_CLARO_HEX = "f8f8f8"
    TEXTO_HEX = "1a1a2e"

    fill_verde = PatternFill("solid", fgColor=VERDE_HEX)
    fill_verde_claro = PatternFill("solid", fgColor=VERDE_CLARO_HEX)
    fill_gris = PatternFill("solid", fgColor=GRIS_CLARO_HEX)
    fill_blanco = PatternFill("solid", fgColor="FFFFFF")

    borde_fino = Border(
        left=Side(style="thin", color="E8E8E8"),
        right=Side(style="thin", color="E8E8E8"),
        top=Side(style="thin", color="E8E8E8"),
        bottom=Side(style="thin", color="E8E8E8"),
    )
    borde_encabezado = Border(
        bottom=Side(style="medium", color=VERDE_OSCURO_HEX),
    )

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Reporte de Desempeño — Empleado #{employee_id}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_verde
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    ws["A2"] = "Generado por PerformTrack · Módulo para MintHCM"
    ws["A2"].font = Font(name="Arial", size=9, color="FFFFFF", italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=VERDE_OSCURO_HEX)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    headers = ["Empleado ID", "KPI", "Valor Actual", "Meta", "% Logro"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = fill_verde
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde_encabezado
    ws.row_dimensions[4].height = 28

    for row_idx, kpi in enumerate(kpis, start=5):
        fila_fill = fill_blanco if (row_idx % 2 == 1) else fill_verde_claro
        valores = [employee_id, kpi.kpi_nombre, kpi.valor_actual, kpi.valor_meta, kpi.porcentaje]

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = Font(name="Arial", size=10, color=TEXTO_HEX)
            cell.fill = fila_fill
            cell.border = borde_fino
            cell.alignment = Alignment(
                horizontal="center" if col_idx != 2 else "left",
                vertical="center"
            )
        ws.row_dimensions[row_idx].height = 22

    anchos = [14, 36, 14, 10, 10]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.freeze_panes = "A5"

    wb.save(output_path)